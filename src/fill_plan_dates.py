"""
Füllt das Plan-Blatt automatisch mit allen Datumszeilen eines Monats vor.
Nutzer muss nur noch Namen + Anteile eintragen.
"""

from pathlib import Path
from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.styles import numbers
import sys


def fill_plan_with_dates(template_path, output_path, year, month):
    """
    Lädt die Vorlage und füllt Spalte A (Datum) im Plan-Blatt
    mit allen Tagen des angegebenen Monats.
    """
    # Validate input parameters
    if not (1 <= month <= 12):
        print(f"❌ Fehler: Ungültiger Monat '{month}'. Monat muss zwischen 1 und 12 liegen")
        return

    if year < 1900 or year > 2100:
        print(f"❌ Fehler: Ungültiges Jahr '{year}'. Jahr muss zwischen 1900 und 2100 liegen")
        return

    # Load template workbook
    try:
        wb = load_workbook(template_path)
    except FileNotFoundError:
        print(f"❌ Fehler: Vorlagendatei '{template_path}' nicht gefunden")
        return
    except PermissionError:
        print(f"❌ Fehler: Keine Berechtigung zum Lesen der Datei '{template_path}'")
        return
    except Exception as e:
        print(f"❌ Fehler beim Laden der Vorlagendatei '{template_path}': {e}")
        return

    try:
        # Regeln-Blatt: Monat_Auswahl setzen
        if "Regeln" in wb.sheetnames:
            regeln_ws = wb["Regeln"]
            # Zeile 7, Spalte B = Monat_Auswahl
            regeln_ws["B7"] = date(year, month, 1)

        # Plan-Blatt füllen
        if "Plan" not in wb.sheetnames:
            print("❌ Blatt 'Plan' nicht gefunden!")
            return

        plan_ws = wb["Plan"]

        # Startdatum
        try:
            start_date = date(year, month, 1)
        except ValueError as e:
            print(f"❌ Fehler: Ungültiges Datum für Jahr {year}, Monat {month}: {e}")
            return

        # Letzter Tag des Monats
        try:
            if month == 12:
                end_date = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(year, month + 1, 1) - timedelta(days=1)
        except ValueError as e:
            print(f"❌ Fehler beim Berechnen des Enddatums: {e}")
            return

        # Alle Tage durchgehen
        current_date = start_date
        row = 2  # Zeile 2 = erste Datenzeile nach Header

        while current_date <= end_date:
            cell = plan_ws[f"A{row}"]
            cell.value = current_date
            cell.number_format = 'DD.MM.YYYY'  # Deutsches Datumsformat
            # Spalten B (Mitarbeiter) und C (Anteil) bleiben leer zum Ausfüllen
            current_date += timedelta(days=1)
            row += 1

        # Save output file
        try:
            wb.save(output_path)
        except PermissionError:
            print(f"❌ Fehler: Keine Berechtigung zum Speichern der Datei '{output_path}'")
            return
        except OSError as e:
            print(f"❌ Fehler beim Speichern der Datei '{output_path}': {e}")
            return

        print(f"✅ Plan-Blatt vorbefüllt für {month:02d}/{year}")
        print(f"   Ausgabe: {output_path}")
        print(f"   Trage jetzt nur noch in Spalte B (Mitarbeiter) und C (Anteil) die Namen ein!")

    except Exception as e:
        print(f"❌ Unerwarteter Fehler beim Füllen des Plan-Blatts: {e}")
        return


if __name__ == "__main__":
    template = Path("templates/Dienstplan_Vorlage_V2_NRW.xlsx")
    
    if len(sys.argv) >= 3:
        year = int(sys.argv[1])
        month = int(sys.argv[2])
    else:
        # Standard: November 2025
        year = 2025
        month = 11
    
    output = Path(f"output/Dienstplan_{year}_{month:02d}_NRW.xlsx")
    output.parent.mkdir(exist_ok=True)
    
    if not template.exists():
        print(f"❌ Vorlage nicht gefunden: {template}")
        print("   Führe erst 'python src/build_template.py' aus!")
        sys.exit(1)
    
    fill_plan_with_dates(template, output, year, month)
