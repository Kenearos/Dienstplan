"""Builds an empty Excel template for the NRW duty roster rules (Variante 2 - streng)."""

from pathlib import Path
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

TEMPLATE_PATH = Path("templates/Dienstplan_Vorlage_V2_NRW.xlsx")
MAX_PLAN_ROWS = 400
MAX_HOLIDAY_ROWS = 50

NRW_HOLIDAYS_2025 = [
    ("2025-01-01", "Neujahr", "NRW"),
    ("2025-04-18", "Karfreitag", "NRW"),
    ("2025-04-21", "Ostermontag", "NRW"),
    ("2025-05-01", "Tag der Arbeit", "NRW"),
    ("2025-05-29", "Christi Himmelfahrt", "NRW"),
    ("2025-06-09", "Pfingstmontag", "NRW"),
    ("2025-06-19", "Fronleichnam", "NRW"),
    ("2025-10-03", "Tag der Deutschen Einheit", "NRW"),
    ("2025-11-01", "Allerheiligen", "NRW"),
    ("2025-12-25", "1. Weihnachtstag", "NRW"),
    ("2025-12-26", "2. Weihnachtstag", "NRW"),
]

NRW_HOLIDAYS_2026 = [
    ("2026-01-01", "Neujahr", "NRW"),
    ("2026-04-03", "Karfreitag", "NRW"),
    ("2026-04-06", "Ostermontag", "NRW"),
    ("2026-05-01", "Tag der Arbeit", "NRW"),
    ("2026-05-14", "Christi Himmelfahrt", "NRW"),
    ("2026-05-25", "Pfingstmontag", "NRW"),
    ("2026-06-04", "Fronleichnam", "NRW"),
    ("2026-10-03", "Tag der Deutschen Einheit", "NRW"),
    ("2026-11-01", "Allerheiligen", "NRW"),
    ("2026-12-25", "1. Weihnachtstag", "NRW"),
    ("2026-12-26", "2. Weihnachtstag", "NRW"),
]


def _style_header(ws, row=1):
    """Apply header styling."""
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[row]:
        if cell.value:
            cell.font = font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")


def _populate_readme(ws):
    ws["A1"] = "NRW-Dienstplan (Variante 2 – streng)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Kurzregeln"
    ws["A3"].font = Font(bold=True)
    rules = [
        "WE-Tag = Fr/Sa/So/Feiertag/Vortag (BL-abhängig).",
        "Variante 2 (streng): WE werden nur vergütet, wenn im Monat ≥ 2,0 WE-Einheiten erreicht werden;",
        "dann 450 €/WE und Abzug 2,0 (Freitag zuerst). WT werden immer mit 250 € vergütet.",
        "Splits anteilig. Monat und Bundesland in 'Regeln' wählen.",
        "",
        "Schritte:",
        "1. In 'Regeln': Monat_Auswahl (erster Tag) + BL_Auswahl setzen.",
        "2. 'Feiertage' kontrollieren bzw. erweitern.",
        "3. Im Blatt 'Plan' pro Tag Datum, Mitarbeiter und Anteil (0–1) eintragen.",
        "4. Auswertung erfolgt automatisch im Blatt 'Auswertung'.",
        "5. 'Checks' zeigt Unstimmigkeiten (Summe Anteil ≠ 1, etc.).",
    ]
    for idx, text in enumerate(rules, start=4):
        ws[f"A{idx}"] = text
    ws.column_dimensions["A"].width = 100


def _populate_rules(ws):
    headers = ["Parameter", "Wert", "Beschreibung"]
    ws.append(headers)
    rows = [
        ("Satz_WT", 250, "Euro für jeden Werktagsdienst (Mo–Do, sofern kein WE-Tag)"),
        ("Satz_WE", 450, "Euro für jeden WE-Tag (Fr–So, Feiertag, Vortag Feiertag)"),
        ("WE_Schwelle", 2.0, "Ab dieser WE-Anzahl wird vergütet (sonst 0 €)"),
        ("Abzug_nach_WE_Schwelle", 2.0, "Einheiten, die nach Erreichen der Schwelle abgezogen werden"),
        ("BL_Auswahl", "NRW", "Bundesland (steuert Feiertage)"),
        ("Monat_Auswahl", date(2025, 11, 1), "Erster Tag des Zielmonats"),
        ("Variante", 2, "Fix: 2 = streng (WE nur bei Schwelle ≥ 2,0)"),
    ]
    for param, value, desc in rows:
        ws.append([param, value, desc])
    
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 80
    _style_header(ws)


def _populate_holidays(ws):
    headers = ["Datum", "Name", "BL"]
    ws.append(headers)
    
    all_holidays = NRW_HOLIDAYS_2025 + NRW_HOLIDAYS_2026
    for iso_date, name, bl in all_holidays:
        # Convert ISO string to date object
        year, month, day = iso_date.split('-')
        date_obj = date(int(year), int(month), int(day))
        ws.append([date_obj, name, bl])
    
    # Create table
    tab = Table(displayName="tblFeiertage", ref=f"A1:C{len(all_holidays)+1}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 8
    _style_header(ws)
    
    # Format column A as date
    for row in range(2, len(all_holidays) + 2):
        ws[f"A{row}"].number_format = 'DD.MM.YYYY'


def _plan_formulas(row: int) -> dict:
    """Return helper-column formulas for Plan sheet (Variante 2)."""
    date_cell = f"A{row}"
    anteil_cell = f"C{row}"
    
    # Holiday range filtered by BL (Non-365 fallback with SUMPRODUCT)
    holiday_check = f'SUMPRODUCT((tblFeiertage[Datum]={date_cell})*(tblFeiertage[BL]=Regeln!$B$6))>0'
    vortag_check = f'SUMPRODUCT((tblFeiertage[Datum]={date_cell}+1)*(tblFeiertage[BL]=Regeln!$B$6))>0'
    
    return {
        "D": f"=IFERROR({holiday_check},FALSE)",  # Ist_FEIERTAG
        "E": f"=IFERROR({vortag_check},FALSE)",   # Ist_VORTAG
        "F": f"=IFERROR(WEEKDAY({date_cell},2)=5,FALSE)",  # Ist_Freitag
        "G": f"=OR($F{row},WEEKDAY({date_cell},2)=6,WEEKDAY({date_cell},2)=7,$D{row},$E{row})",  # Ist_WE_Tag
        "H": f"=NOT($G{row})",  # Ist_WT_Tag
        "I": f"=IF($H{row},{anteil_cell},0)",  # WT_Einheit
        "J": f"=IF(AND($G{row},$F{row}),{anteil_cell},0)",  # WE_Freitag_Einheit
        "K": f"=IF(AND($G{row},NOT($F{row})),{anteil_cell},0)",  # WE_Andere_Einheit
    }


def _populate_plan(ws):
    headers = [
        "Datum", "Mitarbeiter", "Anteil",
        "Ist_FEIERTAG", "Ist_VORTAG", "Ist_Freitag", "Ist_WE_Tag", "Ist_WT_Tag",
        "WT_Einheit", "WE_Freitag_Einheit", "WE_Andere_Einheit"
    ]
    ws.append(headers)
    _style_header(ws)

    for row in range(2, MAX_PLAN_ROWS + 2):
        formulas = _plan_formulas(row)
        for col_letter, formula in formulas.items():
            ws[f"{col_letter}{row}"] = formula

    # Table
    tab = Table(displayName="tblPlan", ref=f"A1:K{MAX_PLAN_ROWS+1}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Data validation for Anteil
    dv = DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"C2:C{MAX_PLAN_ROWS + 1}")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    for col in "DEFGHIJK":
        ws.column_dimensions[col].width = 13
    
    # Format column A as date
    for row in range(2, MAX_PLAN_ROWS + 2):
        ws[f"A{row}"].number_format = 'DD.MM.YYYY'


def _populate_auswertung(ws):
    headers = [
        "Mitarbeiter", "WT_Einheiten", "WE_Freitag", "WE_Andere", "WE_Gesamt",
        "Schwelle_erreicht", "Abzug_gesamt", "Abzug_Freitag", "Abzug_Andere",
        "WE_bezahlt", "Auszahlung_WT", "Auszahlung_WE", "Auszahlung_Gesamt"
    ]
    ws.append(headers)
    _style_header(ws)

    # Manual employee list (user fills column A)
    # Row 2 onwards: formulas reference column A
    
    monat_start = "Regeln!$B$7"
    monat_end = f"EOMONTH({monat_start},0)"
    
    # Create formulas for 50 rows
    for row in range(2, 52):
        name_ref = f"$A{row}"
        
        # Skip if no name
        guard = f'IF({name_ref}="",""'
        
        # WT_Einheiten - using SUMPRODUCT for compatibility
        wt_formula = (
            f'={guard},SUMPRODUCT((tblPlan[Mitarbeiter]={name_ref})*'
            f'(tblPlan[Datum]>={monat_start})*(tblPlan[Datum]<={monat_end})*'
            f'(tblPlan[WT_Einheit])))'
        )
        ws[f"B{row}"] = wt_formula
        
        # WE_Freitag
        we_fri_formula = (
            f'={guard},SUMPRODUCT((tblPlan[Mitarbeiter]={name_ref})*'
            f'(tblPlan[Datum]>={monat_start})*(tblPlan[Datum]<={monat_end})*'
            f'(tblPlan[WE_Freitag_Einheit])))'
        )
        ws[f"C{row}"] = we_fri_formula
        
        # WE_Andere
        we_other_formula = (
            f'={guard},SUMPRODUCT((tblPlan[Mitarbeiter]={name_ref})*'
            f'(tblPlan[Datum]>={monat_start})*(tblPlan[Datum]<={monat_end})*'
            f'(tblPlan[WE_Andere_Einheit])))'
        )
        ws[f"D{row}"] = we_other_formula
        
        # WE_Gesamt
        ws[f"E{row}"] = f'={guard},C{row}+D{row})'
        
        # Schwelle_erreicht
        ws[f"F{row}"] = f'={guard},IF(E{row}>=Regeln!$B$4-0.0001,"JA","NEIN"))'
        
        # Abzug_gesamt
        ws[f"G{row}"] = f'={guard},IF(E{row}>=Regeln!$B$4-0.0001,Regeln!$B$5,0))'
        
        # Abzug_Freitag
        ws[f"H{row}"] = f'={guard},MIN(G{row},C{row}))'
        
        # Abzug_Andere
        ws[f"I{row}"] = f'={guard},MAX(0,G{row}-H{row}))'
        
        # WE_bezahlt (Variante 2: only if threshold reached)
        ws[f"J{row}"] = f'={guard},IF(E{row}<Regeln!$B$4-0.0001,0,(C{row}-H{row})+(D{row}-I{row})))'
        
        # Auszahlung_WT
        ws[f"K{row}"] = f'={guard},B{row}*Regeln!$B$2)'
        
        # Auszahlung_WE
        ws[f"L{row}"] = f'={guard},J{row}*Regeln!$B$3)'
        
        # Auszahlung_Gesamt
        ws[f"M{row}"] = f'={guard},K{row}+L{row})'

    widths = [22, 14, 14, 14, 14, 16, 14, 14, 14, 14, 14, 14, 16]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _populate_checks(ws):
    ws["A1"] = "Datum"
    ws["B1"] = "Summe_Anteile"
    ws["C1"] = "Status"
    _style_header(ws)
    
    # Manual check list - user can add dates to check
    # Formula checks sum of Anteil for each date
    for row in range(2, 52):
        date_ref = f"A{row}"
        ws[f"B{row}"] = f'=IF({date_ref}="","",SUMPRODUCT((tblPlan[Datum]={date_ref})*(tblPlan[Anteil])))'
        ws[f"C{row}"] = f'=IF({date_ref}="","",IF(ABS(B{row}-1)<=0.0001,"OK","FEHLER"))'
    
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12


def build_template():
    TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    readme_ws = wb.active
    readme_ws.title = "README"
    _populate_readme(readme_ws)

    rules_ws = wb.create_sheet("Regeln")
    _populate_rules(rules_ws)

    holiday_ws = wb.create_sheet("Feiertage")
    _populate_holidays(holiday_ws)

    plan_ws = wb.create_sheet("Plan")
    _populate_plan(plan_ws)

    auswertung_ws = wb.create_sheet("Auswertung")
    _populate_auswertung(auswertung_ws)

    checks_ws = wb.create_sheet("Checks")
    _populate_checks(checks_ws)

    wb.save(TEMPLATE_PATH)
    return TEMPLATE_PATH


if __name__ == "__main__":
    path = build_template()
    print(f"✅ Vorlage (Variante 2 – streng) erstellt: {path}")

