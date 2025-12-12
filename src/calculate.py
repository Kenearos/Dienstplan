"""
Berechnet die Vergütung aus der Plan-Datei nach NRW-Regeln (Variante 2)
"""

from pathlib import Path
from datetime import datetime, timedelta, date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import sys
from collections import defaultdict


# Vergütungssätze
SATZ_WT = 250  # Euro für Werktag
SATZ_WE = 450  # Euro für Wochenende
WE_SCHWELLE = 2.0  # Mindestanzahl WE-Dienste für Vergütung
ABZUG = 1.0  # Abzug nach Erreichen der Schwelle


def load_holidays(wb):
    """Lädt Feiertage aus dem Feiertage-Blatt."""
    if "Feiertage" not in wb.sheetnames:
        return set()
    
    holidays = set()
    ws = wb["Feiertage"]
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[2] == "NRW":  # Datum und BL prüfen
            date_raw = row[0]
            if isinstance(date_raw, str):
                try:
                    parsed_date = datetime.strptime(date_raw, '%d.%m.%Y').date()
                    holidays.add(parsed_date)
                except:
                    pass
            elif isinstance(date_raw, datetime):
                holidays.add(date_raw.date())
            elif isinstance(date_raw, date):
                holidays.add(date_raw)
    
    return holidays


def is_we_tag(datum, holidays):
    """Prüft ob ein Datum ein WE-Tag ist (Fr/Sa/So/Feiertag/Vortag)."""
    if isinstance(datum, datetime):
        datum = datum.date()
    
    # Freitag (4), Samstag (5), Sonntag (6)
    weekday = datum.weekday()
    if weekday >= 4:  # Fr, Sa, So
        return True
    
    # Ist Feiertag?
    if datum in holidays:
        return True
    
    # Ist Vortag eines Feiertags?
    next_day = datum + timedelta(days=1)
    if next_day in holidays:
        return True
    
    return False


def is_freitag(datum):
    """Prüft ob ein Datum ein Freitag ist."""
    if isinstance(datum, datetime):
        datum = datum.date()
    return datum.weekday() == 4


def calculate_verguetung(plan_data, holidays):
    """Berechnet Vergütung je Mitarbeiter. Anteil wird automatisch berechnet."""
    
    # Gruppiere nach Datum und zähle Mitarbeiter
    dienste_pro_tag = defaultdict(list)
    for datum, mitarbeiter in plan_data:
        if mitarbeiter:
            dienste_pro_tag[datum].append(mitarbeiter)
    
    # Sammle Daten je Mitarbeiter
    mitarbeiter_data = defaultdict(lambda: {
        'wt_einheiten': 0.0,
        'we_freitag': 0.0,
        'we_andere': 0.0
    })
    
    # Berechne Anteile automatisch
    for datum, mitarbeiter_liste in dienste_pro_tag.items():
        anzahl = len(mitarbeiter_liste)
        anteil = 1.0 / anzahl if anzahl > 0 else 0
        
        for mitarbeiter in mitarbeiter_liste:
            if is_we_tag(datum, holidays):
                if is_freitag(datum):
                    mitarbeiter_data[mitarbeiter]['we_freitag'] += anteil
                else:
                    mitarbeiter_data[mitarbeiter]['we_andere'] += anteil
            else:
                mitarbeiter_data[mitarbeiter]['wt_einheiten'] += anteil
    
    # Berechne Vergütung
    results = []
    
    for mitarbeiter, data in sorted(mitarbeiter_data.items()):
        wt = data['wt_einheiten']
        we_fri = data['we_freitag']
        we_other = data['we_andere']
        we_gesamt = we_fri + we_other
        
        # Schwelle erreicht?
        schwelle_erreicht = we_gesamt >= (WE_SCHWELLE - 0.0001)
        
        if schwelle_erreicht:
            # Abzug von 1.0 WE-Einheit (Freitag zuerst)
            abzug_freitag = min(ABZUG, we_fri)
            abzug_andere = max(0, ABZUG - abzug_freitag)

            # Bezahlte WE-Einheiten
            we_bezahlt = (we_fri - abzug_freitag) + (we_other - abzug_andere)

            # Auszahlungen - nur wenn Schwelle erreicht
            auszahlung_wt = wt * SATZ_WT
            auszahlung_we = we_bezahlt * SATZ_WE
        else:
            # Schwelle nicht erreicht - kein Bonus (weder WT noch WE)
            abzug_freitag = 0
            abzug_andere = 0
            we_bezahlt = 0
            auszahlung_wt = 0
            auszahlung_we = 0

        auszahlung_gesamt = auszahlung_wt + auszahlung_we
        
        results.append({
            'mitarbeiter': mitarbeiter,
            'wt_einheiten': wt,
            'we_freitag': we_fri,
            'we_andere': we_other,
            'we_gesamt': we_gesamt,
            'schwelle_erreicht': 'JA' if schwelle_erreicht else 'NEIN',
            'abzug_freitag': abzug_freitag,
            'abzug_andere': abzug_andere,
            'we_bezahlt': we_bezahlt,
            'auszahlung_wt': auszahlung_wt,
            'auszahlung_we': auszahlung_we,
            'auszahlung_gesamt': auszahlung_gesamt
        })
    
    return results


def process_file(filepath):
    """Verarbeitet die Excel-Datei und schreibt Auswertung."""
    
    wb = load_workbook(filepath)
    
    # Lade Feiertage
    holidays = load_holidays(wb)
    print(f"📅 {len(holidays)} Feiertage geladen")
    
    # Lade Plan-Daten
    if "Plan" not in wb.sheetnames:
        print("❌ Blatt 'Plan' nicht gefunden!")
        return
    
    plan_ws = wb["Plan"]
    plan_data = []
    
    for row in plan_ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Wenn Datum vorhanden
            datum_raw = row[0]
            mitarbeiter = row[1] if len(row) > 1 else None
            
            # Parse Datum (kann String oder date sein)
            if isinstance(datum_raw, str):
                try:
                    datum = datetime.strptime(datum_raw, '%d.%m.%Y').date()
                except:
                    continue
            elif isinstance(datum_raw, datetime):
                datum = datum_raw.date()
            elif isinstance(datum_raw, date):
                datum = datum_raw
            else:
                continue
            
            if mitarbeiter:
                plan_data.append((datum, mitarbeiter))
    
    print(f"📋 {len(plan_data)} Einträge im Plan")
    
    # Berechne Vergütung
    results = calculate_verguetung(plan_data, holidays)
    
    # Schreibe Auswertung
    if "Auswertung" not in wb.sheetnames:
        print("❌ Blatt 'Auswertung' nicht gefunden!")
        return
    
    auswertung_ws = wb["Auswertung"]
    
    # Lösche alte Daten (ab Zeile 2)
    auswertung_ws.delete_rows(2, auswertung_ws.max_row)
    
    # Schreibe neue Daten
    for idx, result in enumerate(results, start=2):
        auswertung_ws[f"A{idx}"] = result['mitarbeiter']
        auswertung_ws[f"B{idx}"] = round(result['wt_einheiten'], 2)
        auswertung_ws[f"C{idx}"] = round(result['we_freitag'], 2)
        auswertung_ws[f"D{idx}"] = round(result['we_andere'], 2)
        auswertung_ws[f"E{idx}"] = round(result['we_gesamt'], 2)
        auswertung_ws[f"F{idx}"] = result['schwelle_erreicht']
        auswertung_ws[f"G{idx}"] = round(result['abzug_freitag'], 2)
        auswertung_ws[f"H{idx}"] = round(result['abzug_andere'], 2)
        auswertung_ws[f"I{idx}"] = round(result['we_bezahlt'], 2)
        auswertung_ws[f"J{idx}"] = round(result['auszahlung_wt'], 2)
        auswertung_ws[f"K{idx}"] = round(result['auszahlung_we'], 2)
        auswertung_ws[f"L{idx}"] = round(result['auszahlung_gesamt'], 2)
        
        # Formatierung für Schwelle
        if result['schwelle_erreicht'] == 'JA':
            auswertung_ws[f"F{idx}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            auswertung_ws[f"F{idx}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    wb.save(filepath)
    print(f"\n✅ Auswertung geschrieben: {len(results)} Mitarbeiter")
    print(f"   Datei: {filepath}")
    
    # Zeige Zusammenfassung
    print(f"\n{'='*70}")
    print(f"{'Mitarbeiter':<20} {'WT':<8} {'WE':<8} {'Schwelle':<10} {'Gesamt':>10}")
    print(f"{'='*70}")
    for r in results:
        print(f"{r['mitarbeiter']:<20} {r['wt_einheiten']:>6.1f}  {r['we_gesamt']:>6.1f}  {r['schwelle_erreicht']:<10} {r['auszahlung_gesamt']:>9.2f} €")
    print(f"{'='*70}")


if __name__ == "__main__":
    if len(sys.argv) >= 2:
        filepath = Path(sys.argv[1])
    else:
        filepath = Path("output/Dienstplan_2025_11_NRW.xlsx")
    
    if not filepath.exists():
        print(f"❌ Datei nicht gefunden: {filepath}")
        sys.exit(1)
    
    process_file(filepath)
