"""
Fügt Testdaten in die Dienstplan-Datei ein
"""

from pathlib import Path
from openpyxl import load_workbook
import sys


def add_test_data(filepath):
    """Fügt Testdaten für November 2025 ein."""
    
    wb = load_workbook(filepath)
    
    if "Plan" not in wb.sheetnames:
        print("❌ Blatt 'Plan' nicht gefunden!")
        return
    
    plan_ws = wb["Plan"]
    
    # Testdaten: Mitarbeiter Namen und Anteile für die ersten Tage
    test_entries = [
        ("Max Mustermann", 1.0),
        ("Anna Schmidt", 1.0),
        ("Max Mustermann", 1.0),
        ("Peter Klein", 0.5),
        ("Anna Schmidt", 0.5),
        ("Max Mustermann", 1.0),
        ("Anna Schmidt", 1.0),
        ("Peter Klein", 1.0),
        ("Max Mustermann", 0.5),
        ("Anna Schmidt", 0.5),
        ("Max Mustermann", 1.0),
        ("Peter Klein", 1.0),
        ("Anna Schmidt", 1.0),
        ("Max Mustermann", 1.0),
        ("Peter Klein", 0.5),
        ("Anna Schmidt", 0.5),
    ]
    
    # Füge die Daten ab Zeile 2 ein
    for i, (name, anteil) in enumerate(test_entries, start=2):
        plan_ws[f"B{i}"] = name
        plan_ws[f"C{i}"] = anteil
    
    # Mitarbeiterliste in Auswertung
    if "Auswertung" in wb.sheetnames:
        auswertung_ws = wb["Auswertung"]
        mitarbeiter = ["Max Mustermann", "Anna Schmidt", "Peter Klein"]
        for i, name in enumerate(mitarbeiter, start=2):
            auswertung_ws[f"A{i}"] = name
    
    wb.save(filepath)
    print(f"✅ Testdaten eingefügt in {filepath}")
    print(f"   Mitarbeiter: Max Mustermann, Anna Schmidt, Peter Klein")
    print(f"   {len(test_entries)} Einträge hinzugefügt")


if __name__ == "__main__":
    if len(sys.argv) >= 2:
        filepath = Path(sys.argv[1])
    else:
        filepath = Path("output/Dienstplan_2025_11_NRW.xlsx")
    
    if not filepath.exists():
        print(f"❌ Datei nicht gefunden: {filepath}")
        sys.exit(1)
    
    add_test_data(filepath)
