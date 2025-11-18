"""
Füllt das Plan-Blatt mit Datumswerten und fügt Testdaten hinzu (vereinfachte Version)
"""

from pathlib import Path
from datetime import date, timedelta
from openpyxl import load_workbook
import sys


def fill_plan_with_test_data(template_path, output_path, year, month):
    """Lädt die Vorlage, füllt Datum und fügt Testdaten ein."""
    
    wb = load_workbook(template_path)
    
    if "Plan" not in wb.sheetnames:
        print("❌ Blatt 'Plan' nicht gefunden!")
        return
    
    plan_ws = wb["Plan"]
    
    # Startdatum
    start_date = date(year, month, 1)
    
    # Letzter Tag des Monats
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)
    
    # Testmitarbeiter
    mitarbeiter = ["Max Mustermann", "Anna Schmidt", "Peter Klein"]
    
    # Alle Tage durchgehen und mit Testdaten füllen
    current_date = start_date
    row = 2
    
    while current_date <= end_date:
        # Jeden 4. Tag: Split zwischen 2 Mitarbeitern
        # Sonst: ein Mitarbeiter
        if (row - 2) % 4 == 0:
            # Split-Tag: 2 Einträge für denselben Tag
            idx1 = (row - 2) % len(mitarbeiter)
            idx2 = (idx1 + 1) % len(mitarbeiter)
            
            # Erster Mitarbeiter
            plan_ws[f"A{row}"] = current_date.strftime('%d.%m.%Y')
            plan_ws[f"B{row}"] = mitarbeiter[idx1]
            row += 1
            
            # Zweiter Mitarbeiter (gleiches Datum)
            plan_ws[f"A{row}"] = current_date.strftime('%d.%m.%Y')
            plan_ws[f"B{row}"] = mitarbeiter[idx2]
            row += 1
        else:
            # Normaler Tag: 1 Mitarbeiter
            idx = (row - 2) % len(mitarbeiter)
            
            plan_ws[f"A{row}"] = current_date.strftime('%d.%m.%Y')
            plan_ws[f"B{row}"] = mitarbeiter[idx]
            row += 1
        
        current_date += timedelta(days=1)
    
    wb.save(output_path)
    print(f"✅ Plan vorbefüllt für {month:02d}/{year}")
    print(f"   Ausgabe: {output_path}")
    print(f"   Testdaten: {len(mitarbeiter)} Mitarbeiter für alle {row-2} Tage")


if __name__ == "__main__":
    template = Path("templates/Dienstplan_Vorlage_V2_NRW_Simple.xlsx")
    
    if len(sys.argv) >= 3:
        year = int(sys.argv[1])
        month = int(sys.argv[2])
    else:
        year = 2025
        month = 11
    
    output = Path(f"output/Dienstplan_{year}_{month:02d}_NRW.xlsx")
    output.parent.mkdir(exist_ok=True)
    
    if not template.exists():
        print(f"❌ Vorlage nicht gefunden: {template}")
        print("   Führe erst 'python src/build_template_simple.py' aus!")
        sys.exit(1)
    
    fill_plan_with_test_data(template, output, year, month)
