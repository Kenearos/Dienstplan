"""Builds a simplified Excel template without complex formulas."""

from pathlib import Path
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.worksheet.datavalidation import DataValidation

TEMPLATE_PATH = Path("templates/Dienstplan_Vorlage_V2_NRW_Simple.xlsx")

NRW_HOLIDAYS_2025 = [
    (date(2025, 1, 1), "Neujahr", "NRW"),
    (date(2025, 4, 18), "Karfreitag", "NRW"),
    (date(2025, 4, 21), "Ostermontag", "NRW"),
    (date(2025, 5, 1), "Tag der Arbeit", "NRW"),
    (date(2025, 5, 29), "Christi Himmelfahrt", "NRW"),
    (date(2025, 6, 9), "Pfingstmontag", "NRW"),
    (date(2025, 6, 19), "Fronleichnam", "NRW"),
    (date(2025, 10, 3), "Tag der Deutschen Einheit", "NRW"),
    (date(2025, 11, 1), "Allerheiligen", "NRW"),
    (date(2025, 12, 25), "1. Weihnachtstag", "NRW"),
    (date(2025, 12, 26), "2. Weihnachtstag", "NRW"),
]

NRW_HOLIDAYS_2026 = [
    (date(2026, 1, 1), "Neujahr", "NRW"),
    (date(2026, 4, 3), "Karfreitag", "NRW"),
    (date(2026, 4, 6), "Ostermontag", "NRW"),
    (date(2026, 5, 1), "Tag der Arbeit", "NRW"),
    (date(2026, 5, 14), "Christi Himmelfahrt", "NRW"),
    (date(2026, 5, 25), "Pfingstmontag", "NRW"),
    (date(2026, 6, 4), "Fronleichnam", "NRW"),
    (date(2026, 10, 3), "Tag der Deutschen Einheit", "NRW"),
    (date(2026, 11, 1), "Allerheiligen", "NRW"),
    (date(2026, 12, 25), "1. Weihnachtstag", "NRW"),
    (date(2026, 12, 26), "2. Weihnachtstag", "NRW"),
]


def _style_header(ws, row=1):
    """Apply header styling."""
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[row]:
        if cell.value:
            cell.font = font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")


def build_simple_template():
    """Creates a simple template without complex formulas."""
    TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    
    # README
    readme_ws = wb.active
    readme_ws.title = "README"
    readme_ws["A1"] = "NRW-Dienstplan - Einfache Version"
    readme_ws["A1"].font = Font(bold=True, size=14)
    readme_ws["A3"] = "Anleitung:"
    readme_ws["A4"] = "1. Trage im Blatt 'Plan' nur Datum und Mitarbeiter ein"
    readme_ws["A5"] = "2. Pro Tag kannst du 1 oder 2 Mitarbeiter eintragen (Split = je 0.5)"
    readme_ws["A6"] = "3. Führe 'python src/calculate.py' aus, um die Vergütung zu berechnen"
    readme_ws["A7"] = "4. Das Ergebnis erscheint im Blatt 'Auswertung'"
    readme_ws.column_dimensions["A"].width = 80
    
    # Feiertage
    holiday_ws = wb.create_sheet("Feiertage")
    holiday_ws.append(["Datum", "Name", "BL"])
    
    all_holidays = NRW_HOLIDAYS_2025 + NRW_HOLIDAYS_2026
    for holiday_date, name, bl in all_holidays:
        # Format date as string for display
        holiday_ws.append([holiday_date.strftime('%d.%m.%Y'), name, bl])
    
    holiday_ws.column_dimensions["A"].width = 14
    holiday_ws.column_dimensions["B"].width = 32
    holiday_ws.column_dimensions["C"].width = 8
    _style_header(holiday_ws)
    
    # Plan (simple input sheet)
    plan_ws = wb.create_sheet("Plan")
    plan_ws.append(["Datum", "Mitarbeiter"])
    _style_header(plan_ws)
    
    plan_ws.column_dimensions["A"].width = 14
    plan_ws.column_dimensions["B"].width = 25
    
    # Auswertung (will be filled by Python script)
    auswertung_ws = wb.create_sheet("Auswertung")
    headers = [
        "Mitarbeiter",
        "WT_Dienste",
        "WE_Dienste_Freitag",
        "WE_Dienste_Andere",
        "WE_Gesamt",
        "Schwelle_erreicht",
        "Abzug_Freitag",
        "Abzug_Andere",
        "WE_bezahlt",
        "Auszahlung_WT",
        "Auszahlung_WE",
        "Auszahlung_Gesamt"
    ]
    auswertung_ws.append(headers)
    _style_header(auswertung_ws)
    
    for col_idx in range(1, len(headers) + 1):
        auswertung_ws.column_dimensions[chr(64 + col_idx)].width = 16
    
    wb.save(TEMPLATE_PATH)
    return TEMPLATE_PATH


if __name__ == "__main__":
    path = build_simple_template()
    print(f"✅ Einfache Vorlage erstellt: {path}")
    print("   Verwende diese mit Python-Berechnungen statt Excel-Formeln")
