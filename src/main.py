"""
Excel XLSX Generator
Erstellt Excel-Dateien mit openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from datetime import datetime


def create_example_excel():
    """Erstellt eine Beispiel-Excel-Datei mit formatierten Daten."""

    try:
        # Neues Workbook erstellen
        wb = Workbook()
        ws = wb.active
        ws.title = "Beispiel"

        # Überschriften hinzufügen
        headers = ["Name", "Alter", "Stadt", "Beruf"]
        ws.append(headers)

        # Überschriften formatieren
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Beispieldaten hinzufügen
        data = [
            ["Max Mustermann", 30, "Berlin", "Entwickler"],
            ["Erika Musterfrau", 28, "München", "Designerin"],
            ["Hans Schmidt", 35, "Hamburg", "Manager"],
            ["Anna Weber", 27, "Köln", "Analyst"],
        ]

        for row in data:
            ws.append(row)

        # Spaltenbreiten anpassen
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        # Ausgabeverzeichnis erstellen
        output_dir = Path("output")
        try:
            output_dir.mkdir(exist_ok=True)
        except PermissionError:
            print(f"❌ Fehler: Keine Berechtigung zum Erstellen des Verzeichnisses '{output_dir}'")
            raise
        except OSError as e:
            print(f"❌ Fehler beim Erstellen des Verzeichnisses '{output_dir}': {e}")
            raise

        # Datei speichern
        output_file = output_dir / f"example_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        try:
            wb.save(output_file)
        except PermissionError:
            print(f"❌ Fehler: Keine Berechtigung zum Speichern der Datei '{output_file}'")
            raise
        except OSError as e:
            print(f"❌ Fehler beim Speichern der Datei '{output_file}': {e}")
            raise

        print(f"Excel-Datei erfolgreich erstellt: {output_file}")
        return output_file

    except Exception as e:
        print(f"❌ Unerwarteter Fehler beim Erstellen der Excel-Datei: {e}")
        raise


if __name__ == "__main__":
    create_example_excel()
