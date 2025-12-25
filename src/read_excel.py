"""
Excel-Datei einlesen und Inhalt anzeigen
"""

from openpyxl import load_workbook
import json
from pathlib import Path


def read_excel_to_dict(filepath):
    """Liest eine Excel-Datei und gibt die Daten als Dictionary zurück."""

    try:
        wb = load_workbook(filepath, data_only=True)
    except FileNotFoundError:
        print(f"❌ Fehler: Datei '{filepath}' nicht gefunden")
        raise
    except PermissionError:
        print(f"❌ Fehler: Keine Berechtigung zum Lesen der Datei '{filepath}'")
        raise
    except Exception as e:
        print(f"❌ Fehler beim Laden der Excel-Datei '{filepath}': {e}")
        raise

    result = {}

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Daten aus dem Sheet lesen
            data = []
            for row in ws.iter_rows(values_only=True):
                # Nur Zeilen mit Inhalt
                if any(cell is not None for cell in row):
                    data.append(list(row))

            result[sheet_name] = data
    except Exception as e:
        print(f"❌ Fehler beim Lesen der Daten aus der Excel-Datei: {e}")
        raise

    return result


def print_excel_content(filepath):
    """Gibt den Inhalt einer Excel-Datei formatiert aus."""

    print(f"\n{'='*60}")
    print(f"Excel-Datei: {filepath}")
    print(f"{'='*60}\n")

    try:
        data = read_excel_to_dict(filepath)
    except Exception:
        # Error already printed in read_excel_to_dict
        raise

    try:
        for sheet_name, rows in data.items():
            print(f"\n📊 Sheet: {sheet_name}")
            print(f"{'-'*60}")

            if not rows:
                print("  (leer)")
                continue

            # Tabelle ausgeben
            for i, row in enumerate(rows, 1):
                row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
                print(f"  {i:3d}: {row_str}")

        print(f"\n{'='*60}\n")

        # Als JSON ausgeben
        print("📄 JSON-Format:")
        try:
            print(json.dumps(data, indent=2, ensure_ascii=False))
        except (TypeError, ValueError) as e:
            print(f"❌ Fehler beim Konvertieren zu JSON: {e}")
            raise
    except Exception as e:
        print(f"❌ Fehler beim Ausgeben der Excel-Daten: {e}")
        raise

    return data


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        # Datei als Argument übergeben
        filepath = sys.argv[1]
    else:
        # Nach neuester Datei im output-Ordner suchen
        output_dir = Path("output")
        excel_files = list(output_dir.glob("*.xlsx"))
        
        if not excel_files:
            print("❌ Keine Excel-Dateien im output-Ordner gefunden!")
            print("Verwendung: python src/read_excel.py <pfad-zur-datei>")
            sys.exit(1)
        
        # Neueste Datei verwenden
        filepath = max(excel_files, key=lambda p: p.stat().st_mtime)
    
    print_excel_content(filepath)
